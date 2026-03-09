#!/usr/bin/env python3
"""
KHI Segment Classifier

Reads KHI article data from an Excel file, cleans HTML content,
sends all articles to Claude API for analysis, and assigns
primary/secondary/tertiary segment_ids based on medical topic relevance.

Usage:
    python classify.py <khi_articles.xlsx> <specialties.xlsx> [--output result.xlsx]

Environment:
    ANTHROPIC_API_KEY - Required. Your Anthropic API key.
"""

import argparse
import json
import os
import sys
from pathlib import Path

import anthropic
import openpyxl
from bs4 import BeautifulSoup


def clean_html(html_content: str) -> str:
    """Remove HTML tags and clean up whitespace from article content."""
    if not html_content:
        return ""
    soup = BeautifulSoup(html_content, "html.parser")
    text = soup.get_text(separator=" ", strip=True)
    # Normalize whitespace
    return " ".join(text.split())


def load_articles(filepath: str) -> list[dict]:
    """Load KHI articles from an Excel file. Expects columns: S no, title, teaser, content."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    # Find column indices (case-insensitive, flexible matching)
    col_map = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_lower = str(h).strip().lower()
        if h_lower in ("s no", "sno", "s_no", "no", "number"):
            col_map["sno"] = i
        elif h_lower == "title":
            col_map["title"] = i
        elif h_lower == "teaser":
            col_map["teaser"] = i
        elif h_lower == "content":
            col_map["content"] = i

    if "title" not in col_map or "content" not in col_map:
        print(f"Error: Could not find 'title' and 'content' columns in {filepath}")
        print(f"Found headers: {headers}")
        sys.exit(1)

    articles = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[col_map["title"]] if col_map.get("title") is not None else None
        content = row[col_map["content"]] if col_map.get("content") is not None else None

        # Only include rows that have a title (rows without titles are formatting artifacts)
        if not title:
            continue

        teaser = row[col_map["teaser"]] if col_map.get("teaser") is not None else None

        articles.append({
            "sno": row[col_map["sno"]] if col_map.get("sno") is not None else None,
            "title": str(title) if title else "",
            "teaser": str(teaser) if teaser else "",
            "content_raw": str(content) if content else "",
            "content_clean": clean_html(str(content)) if content else "",
        })

    wb.close()
    return articles


def load_specialties(filepath: str) -> list[dict]:
    """Load specialty segments from an Excel file. Expects: segment_id, specialty_id, segment_name, total reach."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

    col_map = {}
    for i, h in enumerate(headers):
        if "segment_id" in h:
            col_map["segment_id"] = i
        elif "specialty_id" in h:
            col_map["specialty_id"] = i
        elif "segment_name" in h or "name" in h:
            col_map["segment_name"] = i
        elif "reach" in h:
            col_map["total_reach"] = i

    specialties = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col_map.get("segment_id", 0)] is None:
            continue
        specialties.append({
            "segment_id": int(row[col_map["segment_id"]]),
            "segment_name": str(row[col_map["segment_name"]]),
        })

    wb.close()
    return specialties


def build_prompt(articles: list[dict], specialties: list[dict]) -> str:
    """Build the classification prompt for Claude."""
    # Build article summaries
    article_texts = []
    for i, art in enumerate(articles, 1):
        parts = [f"Article {i}:"]
        if art["title"]:
            parts.append(f"  Title: {art['title']}")
        if art["teaser"]:
            parts.append(f"  Teaser: {art['teaser']}")
        if art["content_clean"]:
            # Truncate very long content to keep prompt manageable
            content = art["content_clean"]
            if len(content) > 3000:
                content = content[:3000] + "... [truncated]"
            parts.append(f"  Content: {content}")
        article_texts.append("\n".join(parts))

    all_articles = "\n\n".join(article_texts)

    # Build specialty list
    spec_lines = []
    for s in sorted(specialties, key=lambda x: x["segment_id"]):
        spec_lines.append(f"  - segment_id: {s['segment_id']}, name: {s['segment_name']}")
    spec_list = "\n".join(spec_lines)

    prompt = f"""You are a medical content classifier for esanum.de, a medical education platform.

You are given all articles from a KHI (Kongress Highlights) — a collection of medical articles published from a specific congress. Your task is to classify the ENTIRE KHI (not individual articles) by assigning the most relevant medical specialty segments.

## Available Segments
{spec_list}

## KHI Articles
{all_articles}

## Task
Analyze ALL articles above as a whole. Based on the medical topics, diseases, treatments, and specialties discussed across the entire KHI, assign:

1. **primary_segment_id** — The single most dominant/relevant specialty for this KHI
2. **secondary_segment_id** — The second most relevant specialty
3. **tertiary_segment_id** — The third most relevant specialty

You MUST choose from the segment_ids listed above. Consider:
- The main medical conditions and diseases discussed
- The target physician audience
- Treatment modalities and therapeutic areas
- The overall congress theme

Respond with ONLY a valid JSON object in this exact format:
{{
  "primary_segment_id": <number>,
  "primary_segment_name": "<name>",
  "primary_reasoning": "<1-2 sentence explanation>",
  "secondary_segment_id": <number>,
  "secondary_segment_name": "<name>",
  "secondary_reasoning": "<1-2 sentence explanation>",
  "tertiary_segment_id": <number>,
  "tertiary_segment_name": "<name>",
  "tertiary_reasoning": "<1-2 sentence explanation>"
}}"""

    return prompt


def classify_khi(articles: list[dict], specialties: list[dict]) -> dict:
    """Send articles to Claude API for classification."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("Error: ANTHROPIC_API_KEY environment variable not set.")
        print("Set it with: export ANTHROPIC_API_KEY='your-key-here'")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)
    prompt = build_prompt(articles, specialties)

    print("Sending articles to Claude for classification...")
    print(f"  Articles: {len(articles)}")
    print(f"  Available segments: {len(specialties)}")

    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=1024,
        messages=[{"role": "user", "content": prompt}],
    )

    response_text = message.content[0].text.strip()

    # Extract JSON from response (handle markdown code blocks)
    if "```json" in response_text:
        response_text = response_text.split("```json")[1].split("```")[0].strip()
    elif "```" in response_text:
        response_text = response_text.split("```")[1].split("```")[0].strip()

    try:
        result = json.loads(response_text)
    except json.JSONDecodeError as e:
        print(f"Error parsing Claude response as JSON: {e}")
        print(f"Raw response:\n{response_text}")
        sys.exit(1)

    # Validate segment_ids exist in our list
    valid_ids = {s["segment_id"] for s in specialties}
    for key in ("primary_segment_id", "secondary_segment_id", "tertiary_segment_id"):
        if result.get(key) not in valid_ids:
            print(f"Warning: {key}={result.get(key)} is not in the allowed segment list.")

    return result


def write_output(
    input_filepath: str,
    output_filepath: str,
    classification: dict,
    articles: list[dict],
):
    """Write a clean output Excel with only real articles + classification columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KHI Classified"

    # Headers: original article columns (cleaned) + classification
    headers = [
        "S no",
        "title",
        "teaser",
        "content",
        "primary_segment_id",
        "primary_segment_name",
        "primary_reasoning",
        "secondary_segment_id",
        "secondary_segment_name",
        "secondary_reasoning",
        "tertiary_segment_id",
        "tertiary_segment_name",
        "tertiary_reasoning",
    ]
    ws.append(headers)

    classification_fields = [
        "primary_segment_id",
        "primary_segment_name",
        "primary_reasoning",
        "secondary_segment_id",
        "secondary_segment_name",
        "secondary_reasoning",
        "tertiary_segment_id",
        "tertiary_segment_name",
        "tertiary_reasoning",
    ]

    # Write each article row with classification (same for all rows since it's KHI-level)
    for art in articles:
        row = [
            art.get("sno", ""),
            art.get("title", ""),
            art.get("teaser", ""),
            art.get("content_clean", ""),
        ]
        for field in classification_fields:
            row.append(classification.get(field, ""))
        ws.append(row)

    # Auto-size a few key columns for readability
    ws.column_dimensions["B"].width = 60  # title
    ws.column_dimensions["C"].width = 40  # teaser

    wb.save(output_filepath)
    wb.close()
    print(f"\nOutput saved to: {output_filepath}")


def main():
    parser = argparse.ArgumentParser(
        description="KHI Segment Classifier - Automatically categorize KHI articles by medical specialty"
    )
    parser.add_argument(
        "khi_file",
        help="Path to the KHI articles Excel file (.xlsx)",
    )
    parser.add_argument(
        "specialties_file",
        help="Path to the specialties/segments Excel file (.xlsx)",
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Output Excel file path (default: <input>_classified.xlsx)",
    )

    args = parser.parse_args()

    # Validate inputs
    if not Path(args.khi_file).exists():
        print(f"Error: KHI file not found: {args.khi_file}")
        sys.exit(1)
    if not Path(args.specialties_file).exists():
        print(f"Error: Specialties file not found: {args.specialties_file}")
        sys.exit(1)

    # Default output path
    if args.output is None:
        stem = Path(args.khi_file).stem
        args.output = str(Path(args.khi_file).parent / f"{stem}_classified.xlsx")

    # Load data
    print(f"Loading articles from: {args.khi_file}")
    articles = load_articles(args.khi_file)
    print(f"  Found {len(articles)} articles")

    print(f"Loading specialties from: {args.specialties_file}")
    specialties = load_specialties(args.specialties_file)
    print(f"  Found {len(specialties)} segments")

    if not articles:
        print("Error: No articles found in the input file.")
        sys.exit(1)

    # Classify
    classification = classify_khi(articles, specialties)

    # Print results
    print("\n" + "=" * 60)
    print("CLASSIFICATION RESULT")
    print("=" * 60)
    print(f"  Primary:   [{classification['primary_segment_id']}] {classification.get('primary_segment_name', '')}")
    print(f"             {classification.get('primary_reasoning', '')}")
    print(f"  Secondary: [{classification['secondary_segment_id']}] {classification.get('secondary_segment_name', '')}")
    print(f"             {classification.get('secondary_reasoning', '')}")
    print(f"  Tertiary:  [{classification['tertiary_segment_id']}] {classification.get('tertiary_segment_name', '')}")
    print(f"             {classification.get('tertiary_reasoning', '')}")
    print("=" * 60)

    # Write output
    write_output(args.khi_file, args.output, classification, articles)


if __name__ == "__main__":
    main()
