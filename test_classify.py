#!/usr/bin/env python3
"""
Tests for the KHI Segment Classifier.

Run with: python -m pytest test_classify.py -v
Or simply: python test_classify.py
"""

import json
import os
import sys
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl

# Ensure the classifier module is importable
sys.path.insert(0, str(Path(__file__).parent))
from classify import (
    build_prompt,
    clean_html,
    load_articles,
    load_specialties,
    write_output,
)

# ---------------------------------------------------------------------------
# Helpers to create test Excel files
# ---------------------------------------------------------------------------

def create_test_khi_xlsx(filepath, articles):
    """Create a minimal KHI articles Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["S no", "title", "teaser", "content"])
    for art in articles:
        ws.append([art.get("sno"), art.get("title"), art.get("teaser"), art.get("content")])
    # Add some empty rows (simulating real data)
    for _ in range(5):
        ws.append([None, None, None, None])
    wb.save(filepath)


def create_test_specialties_xlsx(filepath):
    """Create a minimal specialties Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["segment_id", "specialty_id", "segment_name", "total reach"])
    specs = [
        (19, 54000, "Onkologie / Hämatologie", 4089),
        (15, 207, "Kardiologie", 7146),
        (1, 90, "Allgemeinmedizin", 41691),
        (18, 540, "Neurologie", 6560),
        (14, 198, "Innere Medizin", 30046),
    ]
    for s in specs:
        ws.append(list(s))
    wb.save(filepath)


# ---------------------------------------------------------------------------
# Test: clean_html
# ---------------------------------------------------------------------------

class TestCleanHtml:
    def test_removes_tags(self):
        assert clean_html("<p>Hello <b>world</b></p>") == "Hello world"

    def test_decodes_entities(self):
        assert clean_html("<p>A &amp; B &ndash; C</p>") == "A & B – C"

    def test_handles_nested_tags(self):
        html = "<div><h2>Title</h2><p>Text with <a href='#'>link</a></p></div>"
        result = clean_html(html)
        assert "Title" in result
        assert "link" in result
        assert "<" not in result

    def test_handles_empty_string(self):
        assert clean_html("") == ""

    def test_handles_none(self):
        assert clean_html(None) == ""

    def test_normalizes_whitespace(self):
        assert clean_html("<p>  too   many    spaces  </p>") == "too many spaces"

    def test_preserves_plain_text(self):
        assert clean_html("No HTML here") == "No HTML here"


# ---------------------------------------------------------------------------
# Test: load_articles
# ---------------------------------------------------------------------------

class TestLoadArticles:
    def test_loads_articles_with_titles(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            create_test_khi_xlsx(f.name, [
                {"sno": 1, "title": "Article 1", "teaser": "Teaser 1", "content": "<p>Content 1</p>"},
                {"sno": 2, "title": "Article 2", "teaser": None, "content": "<p>Content 2</p>"},
            ])
            articles = load_articles(f.name)
            os.unlink(f.name)

        assert len(articles) == 2
        assert articles[0]["title"] == "Article 1"
        assert articles[0]["content_clean"] == "Content 1"
        assert articles[1]["teaser"] == ""

    def test_skips_rows_without_title(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            create_test_khi_xlsx(f.name, [
                {"sno": 1, "title": "Real Article", "teaser": "T", "content": "<p>C</p>"},
                {"sno": None, "title": None, "teaser": None, "content": "<p>Orphan content</p>"},
            ])
            articles = load_articles(f.name)
            os.unlink(f.name)

        assert len(articles) == 1
        assert articles[0]["title"] == "Real Article"

    def test_cleans_html_in_content(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            create_test_khi_xlsx(f.name, [
                {"sno": 1, "title": "Art", "teaser": "T",
                 "content": "<h2>Header</h2><p>Paragraph with <strong>bold</strong></p>"},
            ])
            articles = load_articles(f.name)
            os.unlink(f.name)

        assert "<" not in articles[0]["content_clean"]
        assert "Header" in articles[0]["content_clean"]
        assert "bold" in articles[0]["content_clean"]


# ---------------------------------------------------------------------------
# Test: load_specialties
# ---------------------------------------------------------------------------

class TestLoadSpecialties:
    def test_loads_all_specialties(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            create_test_specialties_xlsx(f.name)
            specs = load_specialties(f.name)
            os.unlink(f.name)

        assert len(specs) == 5
        names = {s["segment_name"] for s in specs}
        assert "Onkologie / Hämatologie" in names
        assert "Kardiologie" in names

    def test_segment_ids_are_ints(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            create_test_specialties_xlsx(f.name)
            specs = load_specialties(f.name)
            os.unlink(f.name)

        for s in specs:
            assert isinstance(s["segment_id"], int)


# ---------------------------------------------------------------------------
# Test: build_prompt
# ---------------------------------------------------------------------------

class TestBuildPrompt:
    def test_contains_all_articles(self):
        articles = [
            {"title": "Oncology Update", "teaser": "New findings", "content_clean": "Cancer treatment advances"},
            {"title": "Cardiology News", "teaser": "Heart health", "content_clean": "Cardiac surgery results"},
        ]
        specs = [
            {"segment_id": 19, "segment_name": "Onkologie"},
            {"segment_id": 15, "segment_name": "Kardiologie"},
        ]
        prompt = build_prompt(articles, specs)

        assert "Oncology Update" in prompt
        assert "Cardiology News" in prompt
        assert "Cancer treatment advances" in prompt

    def test_contains_all_segments(self):
        articles = [{"title": "T", "teaser": "", "content_clean": "C"}]
        specs = [
            {"segment_id": 1, "segment_name": "Allgemeinmedizin"},
            {"segment_id": 19, "segment_name": "Onkologie"},
        ]
        prompt = build_prompt(articles, specs)

        assert "segment_id: 1" in prompt
        assert "Allgemeinmedizin" in prompt
        assert "segment_id: 19" in prompt

    def test_requests_json_output(self):
        articles = [{"title": "T", "teaser": "", "content_clean": "C"}]
        specs = [{"segment_id": 1, "segment_name": "Test"}]
        prompt = build_prompt(articles, specs)

        assert "primary_segment_id" in prompt
        assert "secondary_segment_id" in prompt
        assert "tertiary_segment_id" in prompt
        assert "JSON" in prompt

    def test_truncates_long_content(self):
        articles = [{"title": "T", "teaser": "", "content_clean": "x" * 5000}]
        specs = [{"segment_id": 1, "segment_name": "Test"}]
        prompt = build_prompt(articles, specs)

        assert "[truncated]" in prompt


# ---------------------------------------------------------------------------
# Test: write_output
# ---------------------------------------------------------------------------

class TestWriteOutput:
    def test_creates_clean_output(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as inp:
            create_test_khi_xlsx(inp.name, [
                {"sno": 1, "title": "Art 1", "teaser": "T", "content": "<p>C</p>"},
                {"sno": 2, "title": "Art 2", "teaser": "T2", "content": "<p>C2</p>"},
            ])
            output_path = inp.name.replace(".xlsx", "_out.xlsx")

            classification = {
                "primary_segment_id": 19,
                "primary_segment_name": "Onkologie / Hämatologie",
                "primary_reasoning": "Cancer congress",
                "secondary_segment_id": 15,
                "secondary_segment_name": "Kardiologie",
                "secondary_reasoning": "Some cardio",
                "tertiary_segment_id": 1,
                "tertiary_segment_name": "Allgemeinmedizin",
                "tertiary_reasoning": "General",
            }
            articles = [
                {"sno": 1, "title": "Art 1", "teaser": "T", "content_raw": "<p>C</p>", "content_clean": "C"},
                {"sno": 2, "title": "Art 2", "teaser": "T2", "content_raw": "<p>C2</p>", "content_clean": "C2"},
            ]

            write_output(inp.name, output_path, classification, articles)

            wb = openpyxl.load_workbook(output_path)
            ws = wb.active

            # Should have header + 2 article rows (no empty rows)
            assert ws.max_row == 3

            # Check headers
            headers = [cell.value for cell in ws[1] if cell.value]
            assert "primary_segment_id" in headers
            assert "secondary_segment_id" in headers
            assert "tertiary_segment_id" in headers
            assert "content" in headers
            assert "title" in headers

            # Check classification values on each row
            pid_col = headers.index("primary_segment_id") + 1
            assert ws.cell(row=2, column=pid_col).value == 19
            assert ws.cell(row=3, column=pid_col).value == 19

            # Check article data
            title_col = headers.index("title") + 1
            assert ws.cell(row=2, column=title_col).value == "Art 1"
            assert ws.cell(row=3, column=title_col).value == "Art 2"

            # Content column should have cleaned text, no HTML
            content_col = headers.index("content") + 1
            assert ws.cell(row=2, column=content_col).value == "C"
            assert ws.cell(row=3, column=content_col).value == "C2"

            wb.close()
            os.unlink(inp.name)
            os.unlink(output_path)


# ---------------------------------------------------------------------------
# Test: classify_khi (mocked API call)
# ---------------------------------------------------------------------------

class TestClassifyKhi:
    @patch("classify.anthropic.Anthropic")
    def test_parses_valid_response(self, mock_anthropic_cls):
        os.environ["ANTHROPIC_API_KEY"] = "test-key"

        mock_response = json.dumps({
            "primary_segment_id": 19,
            "primary_segment_name": "Onkologie / Hämatologie",
            "primary_reasoning": "Cancer focus",
            "secondary_segment_id": 11,
            "secondary_segment_name": "Frauenheilkunde",
            "secondary_reasoning": "Gynecologic oncology",
            "tertiary_segment_id": 14,
            "tertiary_segment_name": "Innere Medizin",
            "tertiary_reasoning": "Internal medicine overlap",
        })

        mock_message = MagicMock()
        mock_message.content = [MagicMock(text=mock_response)]
        mock_client = MagicMock()
        mock_client.messages.create.return_value = mock_message
        mock_anthropic_cls.return_value = mock_client

        from classify import classify_khi

        articles = [{"title": "Test", "teaser": "", "content_clean": "Oncology content"}]
        specs = [{"segment_id": 19, "segment_name": "Onkologie / Hämatologie"}]

        result = classify_khi(articles, specs)
        assert result["primary_segment_id"] == 19
        assert result["secondary_segment_id"] == 11

    @patch("classify.anthropic.Anthropic")
    def test_handles_markdown_wrapped_json(self, mock_anthropic_cls):
        os.environ["ANTHROPIC_API_KEY"] = "test-key"

        mock_response = '```json\n{"primary_segment_id": 19, "primary_segment_name": "Onko", "primary_reasoning": "r", "secondary_segment_id": 11, "secondary_segment_name": "Frau", "secondary_reasoning": "r", "tertiary_segment_id": 14, "tertiary_segment_name": "Inn", "tertiary_reasoning": "r"}\n```'

        mock_message = MagicMock()
        mock_message.content = [MagicMock(text=mock_response)]
        mock_client = MagicMock()
        mock_client.messages.create.return_value = mock_message
        mock_anthropic_cls.return_value = mock_client

        from classify import classify_khi

        result = classify_khi(
            [{"title": "T", "teaser": "", "content_clean": "C"}],
            [{"segment_id": 19, "segment_name": "Onko"}],
        )
        assert result["primary_segment_id"] == 19


# ---------------------------------------------------------------------------
# Test: end-to-end with real files (if API key is set)
# ---------------------------------------------------------------------------

class TestEndToEnd:
    def test_with_real_api(self):
        """End-to-end test using the actual Claude API and sample data files."""
        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key or api_key == "test-key":
            print("SKIP: ANTHROPIC_API_KEY not set, skipping end-to-end test")
            return

        khi_path = Path(__file__).parent / "data/khi_content_for_Ai.xlsx"
        spec_path = Path(__file__).parent / "data/main_specialties.xlsx"

        if not khi_path.exists() or not spec_path.exists():
            print("SKIP: Sample data files not found, skipping end-to-end test")
            return

        from classify import classify_khi, load_articles, load_specialties

        articles = load_articles(str(khi_path))
        specialties = load_specialties(str(spec_path))

        assert len(articles) == 17
        assert len(specialties) == 25

        result = classify_khi(articles, specialties)

        # Validate structure
        assert "primary_segment_id" in result
        assert "secondary_segment_id" in result
        assert "tertiary_segment_id" in result

        # Validate segment_ids are from the allowed list
        valid_ids = {s["segment_id"] for s in specialties}
        assert result["primary_segment_id"] in valid_ids, f"Primary {result['primary_segment_id']} not in valid IDs"
        assert result["secondary_segment_id"] in valid_ids, f"Secondary {result['secondary_segment_id']} not in valid IDs"
        assert result["tertiary_segment_id"] in valid_ids, f"Tertiary {result['tertiary_segment_id']} not in valid IDs"

        # All three should be different
        ids = [result["primary_segment_id"], result["secondary_segment_id"], result["tertiary_segment_id"]]
        assert len(set(ids)) == 3, f"Expected 3 distinct segment_ids, got {ids}"

        # For this specific KHI (Deutscher Krebskongress), primary should be Oncology
        assert result["primary_segment_id"] == 19, (
            f"Expected primary=19 (Onkologie) for cancer congress, got {result['primary_segment_id']}"
        )

        # Reasoning fields should be non-empty
        assert len(result.get("primary_reasoning", "")) > 10
        assert len(result.get("secondary_reasoning", "")) > 10
        assert len(result.get("tertiary_reasoning", "")) > 10

        print(f"END-TO-END PASS: primary={result['primary_segment_id']}, "
              f"secondary={result['secondary_segment_id']}, "
              f"tertiary={result['tertiary_segment_id']}")


# ---------------------------------------------------------------------------
# Test: Web Dashboard
# ---------------------------------------------------------------------------

class TestWebApp:
    def _get_client(self):
        from app import app
        app.config["TESTING"] = True
        return app.test_client()

    def test_index_loads(self):
        client = self._get_client()
        resp = client.get("/")
        assert resp.status_code == 200
        assert b"KHI Segment Classifier" in resp.data

    def test_index_shows_empty_state(self):
        client = self._get_client()
        resp = client.get("/")
        assert b"No classification yet" in resp.data

    def test_classify_rejects_no_file(self):
        client = self._get_client()
        resp = client.post("/classify", data={}, follow_redirects=True)
        assert resp.status_code == 200

    def test_classify_rejects_non_xlsx(self):
        client = self._get_client()
        from io import BytesIO
        data = {"khi_file": (BytesIO(b"not excel"), "test.txt")}
        resp = client.post("/classify", data=data, content_type="multipart/form-data", follow_redirects=True)
        assert b"valid .xlsx" in resp.data

    @patch("app.classify_khi")
    def test_classify_success(self, mock_classify):
        mock_classify.return_value = {
            "primary_segment_id": 19,
            "primary_segment_name": "Onkologie / Hämatologie",
            "primary_reasoning": "Cancer focus",
            "secondary_segment_id": 11,
            "secondary_segment_name": "Frauenheilkunde",
            "secondary_reasoning": "Gynecologic oncology",
            "tertiary_segment_id": 14,
            "tertiary_segment_name": "Innere Medizin",
            "tertiary_reasoning": "Internal medicine",
        }
        os.environ["ANTHROPIC_API_KEY"] = "test-key"

        client = self._get_client()

        # Create a test xlsx in memory
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["S no", "title", "teaser", "content"])
        ws.append([1, "Test Article", "Teaser", "<p>Content</p>"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        data = {"khi_file": (buf, "test_khi.xlsx")}
        resp = client.post("/classify", data=data, content_type="multipart/form-data", follow_redirects=True)

        assert resp.status_code == 200
        assert b"Onkologie" in resp.data
        assert b"Classification Results" in resp.data
        assert b"Test Article" in resp.data

    def test_download_without_result(self):
        client = self._get_client()
        resp = client.get("/download", follow_redirects=True)
        assert resp.status_code == 200

    @patch("app.classify_khi")
    def test_api_endpoint(self, mock_classify):
        mock_classify.return_value = {
            "primary_segment_id": 19,
            "primary_segment_name": "Onkologie",
            "primary_reasoning": "r",
            "secondary_segment_id": 11,
            "secondary_segment_name": "Frau",
            "secondary_reasoning": "r",
            "tertiary_segment_id": 14,
            "tertiary_segment_name": "Inn",
            "tertiary_reasoning": "r",
        }
        os.environ["ANTHROPIC_API_KEY"] = "test-key"

        client = self._get_client()

        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["S no", "title", "teaser", "content"])
        ws.append([1, "Art", "T", "<p>C</p>"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = client.post("/api/classify", data={"khi_file": (buf, "t.xlsx")}, content_type="multipart/form-data")
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["classification"]["primary_segment_id"] == 19
        assert data["num_articles"] == 1


# ---------------------------------------------------------------------------
# Run tests
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import pytest
    sys.exit(pytest.main([__file__, "-v"]))
